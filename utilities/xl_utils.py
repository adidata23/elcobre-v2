import openpyxl

path = "D:\Document\Document for Test\datatestcms.xlsx"


def getRowCount(sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.max_row


def getColumnCount(sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.max_column


def readData(sheetName, rowNumber, columnNumber):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row = rowNumber, column = columnNumber).value


def writeData(sheetName, rowNumber, columnNumber, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row = rowNumber, column = columnNumber).value = data
    workbook.save(path)
